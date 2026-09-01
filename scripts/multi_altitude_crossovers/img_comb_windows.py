# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "pandas", "pyarrow", "numpy", "openpyxl", "h5py", "scipy", "fsspec", "aiohttp", "requests",
# ]
# ///
"""Per-segment surface-validity windows from OPR param spreadsheets + posted img_01 gates.

For each (season, segment) in the multi-altitude crossover model table, derive the
window of surface TWTT for which the combined CSARP_standard product's surface sample
comes from image 1, unblanked (see OPR_Toolbox_Guide.md "img_comb fields"):

    surf_valid_min = max(T_blank, t_start(img1))
    surf_valid_max = T_end(img1) - T_guard

where [T_comb, T_blank, T_guard] is the first triplet of param.array.img_comb (array
worksheet of the season's rds param spreadsheet, gitlab.com/openpolarradar/opr_params)
and [t_start, T_end] is image 1's pulse-compressed time gate, read directly (in product
time) from the Time axis of one posted Data_img_01 file per segment on
data.cresis.ku.edu. Single-image segments fall back to the combined file's record gate.

Outputs: outputs/multi_altitude_crossovers/img_comb_windows.csv
Network reads are cached in outputs/multi_altitude_crossovers/cache_img_comb/.

Usage: uv run scripts/multi_altitude_crossovers/img_comb_windows.py
"""

import io
import json
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
import requests

OUT_DIR = Path("outputs/multi_altitude_crossovers")
CACHE_DIR = OUT_DIR / "cache_img_comb"
PARAMS_DIR = CACHE_DIR / "param_sheets"
GATE_CACHE = CACHE_DIR / "img01_gates.json"
CSV_OUT = OUT_DIR / "img_comb_windows.csv"

PARAMS_RAW = ("https://gitlab.com/api/v4/projects/openpolarradar%2Fopr_params"
              "/repository/files/rds_param_{season}.xlsx/raw?ref=HEAD")
DATA_BASE = "https://data.cresis.ku.edu/data/rds/{season}/CSARP_standard/{seg}/{fn}"


def get_param_sheet(season: str) -> Path:
    PARAMS_DIR.mkdir(parents=True, exist_ok=True)
    p = PARAMS_DIR / f"rds_param_{season}.xlsx"
    if not p.exists():
        r = requests.get(PARAMS_RAW.format(season=season), timeout=120)
        r.raise_for_status()
        p.write_bytes(r.content)
    return p


def parse_matlab_vector(s):
    """'[3e-06 -Inf 1e-06 ...]' -> list of floats (or None)."""
    if s is None or str(s).strip() in ("", "[]", "None"):
        return None
    toks = re.split(r"[\s,]+", str(s).strip().strip("[]").strip())
    out = []
    for t in toks:
        if not t:
            continue
        tl = t.lower()
        out.append(-math.inf if tl in ("-inf",) else (math.inf if tl == "inf" else float(t)))
    return out or None


def n_images_from_imgs(s):
    """Count top-level [..] cells inside the {..} imgs string; None if blank."""
    if s is None or str(s).strip() in ("", "{}", "None"):
        return None
    txt = str(s)
    depth = 0
    n = 0
    for ch in txt:
        if ch == "[":
            if depth == 0:
                n += 1
            depth += 1
        elif ch == "]":
            depth -= 1
    return n or None


def img1_wf_indices(s):
    """Waveform indices making up image 1, from the imgs cell string."""
    if s is None or str(s).strip() in ("", "{}", "None"):
        return set()
    txt = str(s)
    # first top-level [..] cell
    depth = 0
    start = end = None
    for i, ch in enumerate(txt):
        if ch == "[":
            if depth == 0 and start is None:
                start = i
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0 and start is not None:
                end = i
                break
    if start is None or end is None:
        return set()
    cell = txt[start + 1:end]
    wfs = set()
    for m in re.finditer(r"(\d+)\s*\*\s*ones", cell):
        wfs.add(int(m.group(1)))
    if re.search(r"(?<![\d*\s])ones\(|^\s*ones\(", cell):
        wfs.add(1)
    for row in cell.split(";"):
        m = re.match(r"\s*(\d+)\s+\d+", row)
        if m:
            wfs.add(int(m.group(1)))
    return wfs


def load_array_worksheet(season: str) -> dict:
    """segment 'YYYYMMDD_SS' -> img_comb, n_images, and max Tpd of image-1 wfs."""
    import openpyxl

    wb = openpyxl.load_workbook(get_param_sheet(season), read_only=True)
    ws = wb["array"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    ic = hdr.index("img_comb")
    im = hdr.index("imgs")
    # per-wf Tpd columns from the radar worksheet
    rws = wb["radar"]
    rrows = list(rws.iter_rows(values_only=True))
    rhdr, rtyp = rrows[0], rrows[1]
    tpd_cols = {}  # wf index -> column
    for i in range(len(rhdr)):
        if rhdr[i] == "Tpd" and rtyp[i]:
            m = re.search(r"wfs\((\d+)\)", str(rtyp[i]))
            if m:
                tpd_cols[int(m.group(1))] = i
    tpd_by_seg = {}
    for r in rrows[2:]:
        if r[0] is None or r[1] is None:
            continue
        seg = f"{int(r[0]):08d}_{int(r[1]):02d}"
        tpd_by_seg[seg] = {wf: r[i] for wf, i in tpd_cols.items()
                           if r[i] is not None and str(r[i]).strip() != ""}
    out = {}
    for r in rows[2:]:
        if r[0] is None or r[1] is None:
            continue
        seg = f"{int(r[0]):08d}_{int(r[1]):02d}"
        wfs1 = img1_wf_indices(r[im]) or {1}
        tpds = tpd_by_seg.get(seg, {})
        tpd1 = max((float(tpds[w]) for w in wfs1 if w in tpds), default=float("nan"))
        out[seg] = {
            "img_comb": parse_matlab_vector(r[ic]),
            "n_images": n_images_from_imgs(r[im]),
            "tpd_img1": tpd1,
        }
    return out


def _read_time_httprange(url: str):
    """Read the Time vector of a posted echogram (v7.3 via h5py range reads,
    v5 via scipy on a seekable fsspec file). Returns np.array or raises."""
    import fsspec
    import h5py

    f = fsspec.open(url, "rb", block_size=256 * 1024).open()
    try:
        try:
            h = h5py.File(f, "r")
            return np.atleast_1d(np.squeeze(h["Time"][()]))
        except OSError:
            f.seek(0)
            import scipy.io as sio
            m = sio.loadmat(f, variable_names=["Time"], squeeze_me=True)
            return np.atleast_1d(m["Time"])
    finally:
        f.close()


def img1_gate(season: str, seg: str, frame: str, cache: dict):
    """(t_start, t_end, source) of image 1 in product time for one segment.

    Tries the Data_img_01 file for the given frame; falls back to the combined
    file (single-image segments). Returns (nan, nan, 'missing') on failure.
    """
    key = f"{season}/{seg}"
    if key in cache:
        return tuple(cache[key])
    for fn, src in ((f"Data_img_01_{frame}.mat", "img_01"),
                    (f"Data_{frame}.mat", "combined(single-image)")):
        url = DATA_BASE.format(season=season, seg=seg, fn=fn)
        try:
            head = requests.head(url, timeout=60)
            if head.status_code != 200:
                continue
            t = _read_time_httprange(url)
            cache[key] = (float(t[0]), float(t[-1]), src)
            return cache[key]
        except Exception as e:  # noqa: BLE001
            print(f"  WARN {season} {seg}: {fn}: {type(e).__name__}: {e}")
    cache[key] = (float("nan"), float("nan"), "missing")
    return cache[key]


def main():
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache = json.loads(GATE_CACHE.read_text()) if GATE_CACHE.exists() else {}

    df = pd.read_parquet(OUT_DIR / "model_table.parquet", columns=["season", "frame_id"])
    df["seg"] = df.frame_id.str.extract(r"Data_(\d{8}_\d{2})")[0]
    df["frm"] = df.frame_id.str.extract(r"Data_(\d{8}_\d{2}_\d{3})")[0]
    pairs = df.groupby(["season", "seg"])["frm"].first().reset_index()

    sheets = {}
    rows = []
    for i, (season, seg, frm) in enumerate(pairs.itertuples(index=False)):
        if season not in sheets:
            sheets[season] = load_array_worksheet(season)
        arr = sheets[season].get(seg)
        notes = []
        if arr is None:
            notes.append("segment missing from array worksheet")
            img_comb, n_img, tpd1 = None, None, float("nan")
        else:
            img_comb, n_img, tpd1 = arr["img_comb"], arr["n_images"], arr["tpd_img1"]
        if img_comb is not None and n_img is None:
            n_img = len(img_comb) // 3 + 1
        if img_comb is None and n_img is None:
            n_img = 1

        t_start, t_end, src = img1_gate(season, seg, frm, cache)
        if i % 20 == 0:
            GATE_CACHE.write_text(json.dumps(cache))
        if src == "missing":
            notes.append("no img_01/combined file readable; gate NaN")

        if img_comb is not None and n_img and n_img > 1:
            t_comb, t_blank, t_guard = img_comb[0], img_comb[1], img_comb[2]
            vmin = max(t_blank, t_start) if np.isfinite(t_start) else t_blank
            vmax = t_end - t_guard
            source = f"array.img_comb + {src} Time gate"
        elif n_img and n_img > 1:
            # multiple images but no combining recorded: surface still read from
            # combined product; treat image-1 gate as the window, flag it
            t_comb = t_blank = t_guard = float("nan")
            vmin, vmax = t_start, t_end
            source = f"{src} Time gate (img_comb blank)"
            notes.append("multi-image but img_comb blank in sheet")
        else:
            t_comb = t_blank = t_guard = float("nan")
            vmin, vmax = t_start, t_end
            source = f"{src} Time gate (single image)"
            notes.append("single image: window = record gate")

        # Conservative variant: the matched filter loses support within Tpd of the
        # gate end (validated: 2014_Greenland_P3 cliff at ~5-6 us = tend - Tpd(3us),
        # not tend - T_guard(1us)); near the gate start / transmit event the
        # receiver is recovering for ~Tpd.
        if img_comb is not None and n_img and n_img > 1:
            guard_eff = np.nanmax([t_guard, tpd1]) if np.isfinite(tpd1) else t_guard
            vmin_safe = max(v for v in (t_blank, t_start, tpd1) if np.isfinite(v))
            vmax_safe = t_end - guard_eff
        else:
            # single image (or no img_comb): qlook/array combine already trims
            # pulse-compression rolloff from the posted gate; keep the plain window
            vmin_safe, vmax_safe = vmin, vmax

        rows.append({
            "season": season,
            "segment": f"Data_{seg}",
            "surf_valid_min_s": vmin,
            "surf_valid_max_s": vmax,
            "surf_valid_min_safe_s": vmin_safe,
            "surf_valid_max_safe_s": vmax_safe,
            "t_blank_s": t_blank,
            "t_guard_s": t_guard,
            "t_comb_s": t_comb,
            "img1_tend_s": t_end,
            "tpd_img1_s": tpd1,
            "n_images": n_img,
            "source": source,
            "notes": "; ".join(notes),
        })
        print(f"[{i+1}/{len(pairs)}] {season} {seg}: "
              f"valid {rows[-1]['surf_valid_min_s']*1e6 if np.isfinite(rows[-1]['surf_valid_min_s']) else float('nan'):.1f}"
              f"..{rows[-1]['surf_valid_max_s']*1e6 if np.isfinite(rows[-1]['surf_valid_max_s']) else float('nan'):.1f} us ({src})")

    GATE_CACHE.write_text(json.dumps(cache))
    out = pd.DataFrame(rows)
    out.to_csv(CSV_OUT, index=False)
    print(f"\nwrote {CSV_OUT} ({len(out)} rows)")
    with pd.option_context("display.width", 200):
        summ = out.assign(vmax_us=out.surf_valid_max_s * 1e6,
                          vmin_us=out.surf_valid_min_s * 1e6)
        print(summ.groupby("season")[["vmin_us", "vmax_us"]].agg(["min", "median", "max"]).round(1))


if __name__ == "__main__":
    main()
